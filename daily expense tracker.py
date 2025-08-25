from datetime import datetime
from tkinter import *
from tkinter.ttk import Treeview, Style, Combobox
from tkinter import messagebox, Toplevel
from tkcalendar import DateEntry
import uuid
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from tkinter import filedialog

# ----------------------------backend---------------------------------------------

FILENAME = "depenses.csv"

categories = ["Alimentation", "Transport", "Eau et électricité", "Santé", "Loisirs", "Éducation"]

class Expense:
    def __init__(self, amount, category, date, id, description):
        if amount <= 0:
            raise ValueError("Le montant doit être positif")
        if category not in categories:
            raise ValueError("Catégorie non trouvée")
        self.amount = amount
        self.category = category
        self.date = date
        self.id = id
        self.description = description

    def to_dict(self):
        return {
            "amount": self.amount,
            "category": self.category,
            "date": self.date,
            "id": self.id,
            "description": self.description
        }

class ExpenseManager:
    def __init__(self):
        self.expenses = []

    def add_expense(self, expense):
        if not isinstance(expense, Expense):
            raise TypeError("Objet Expense attendu")
        for existing_expense in self.expenses:
            if existing_expense.id == expense.id:
                raise ValueError("ID dupliqué")
        self.expenses.append(expense)

    def remove_expense(self, id):
        for searched_expense in self.expenses:
            if searched_expense.id == id:
                self.expenses.remove(searched_expense)
                return
        raise ValueError("Dépense non trouvée")

    def daily_total(self, given_date):
        total = 0
        for expense in self.expenses:
            expense_date = datetime.strptime(expense.date, "%Y-%m-%d")
            given_date_dt = datetime.strptime(given_date, "%Y-%m-%d")
            if expense_date.date() == given_date_dt.date():
                total += expense.amount
        return total

    def monthly_total(self, year, month):
        total = 0
        for expense in self.expenses:
            expense_date = datetime.strptime(expense.date, "%Y-%m-%d")
            if expense_date.year == year and expense_date.month == month:
                total += expense.amount
        return total

def save_expenses(data):
    with open(FILENAME, "w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["ID", "Date", "Catégorie", "Montant", "Description"])
        writer.writerows(data)

def load_expenses():
    if not os.path.exists(FILENAME):
        return []
    with open(FILENAME, "r", encoding="utf-8") as file:
        reader = csv.reader(file)
        next(reader, None)
        rows = [row for row in reader if row and len(row) >= 5]
    return rows


# ------------------- frontend ---------------------------------------------
class ExpenseApp:
    def __init__(self):
        self.manager = ExpenseManager()
        self.root = Tk()
        self.root.title("Suivi des Dépenses Quotidiennes")
        self.root.state("zoomed")
        self.root.minsize(1000, 600)
        self.root.configure(bg="white")


        self.total_var = StringVar()
        self.daily_total_var = StringVar()
        self.monthly_total_var = StringVar()


        for i in range(6):
            self.root.grid_rowconfigure(i, weight=1)
        for i in range(4):
            self.root.grid_columnconfigure(i, weight=1)

        Label(self.root, text="Suivi des Dépenses Quotidiennes", font=("Segoe UI", 28, "bold"),
              bg="white", fg="black").grid(row=0, column=0, columnspan=4, pady=15)

     
        inputs_frame = Frame(self.root, bg="lightgrey", pady=10, padx=10)
        inputs_frame.grid(row=1, column=0, columnspan=4, sticky="nsew", padx=10, pady=5)

        inputs_frame.grid_columnconfigure(0, weight=0, minsize=120)
        inputs_frame.grid_columnconfigure(1, weight=1)
        inputs_frame.grid_columnconfigure(2, weight=0, minsize=120)
        inputs_frame.grid_columnconfigure(3, weight=1)

        label_font = ("Segoe UI", 12)
        entry_font = ("Segoe UI", 12)


        Label(inputs_frame, text="Revenu", font=("Segoe UI", 12, "bold"), bg="lightgrey").grid(row=0, column=0, padx=10,
                                                                                           pady=5, sticky="w")
        self.revenu_var = StringVar()
        self.revenu_entry = Entry(inputs_frame, font=entry_font, textvariable=self.revenu_var)
        self.revenu_entry.grid(row=0, column=1, padx=10, pady=5, sticky="ew", ipady=5)

        self.revenu_entry.bind("<KeyRelease>", lambda e: self.update_total())
        self.revenu_entry.bind("<FocusOut>", lambda e: self._normalize_and_save_revenu())
        self.revenu_entry.bind("<Return>", lambda e: self._normalize_and_save_revenu())

        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        Label(inputs_frame, text="Reste", font=("Segoe UI", 12, "bold"), bg="lightgrey").grid(row=0, column=2, padx=10,
                                                                                        pady=5, sticky="w")
        self.reste_var = StringVar()
        self.reste_entry = Entry(inputs_frame, font=entry_font, textvariable=self.reste_var, state="readonly")
        self.reste_entry.grid(row=0, column=3, padx=10, pady=5, sticky="ew", ipady=5)


        Label(inputs_frame, text="Montant", font=("Segoe UI", 12, "bold"), bg="lightgrey").grid(row=1, column=0, padx=10, pady=5,
                                                                                  sticky="w")
        self.amount_entry = Entry(inputs_frame, font=entry_font)
        self.amount_entry.grid(row=1, column=1, padx=10, pady=5, sticky="ew", ipady=5)



        Label(inputs_frame, text="Catégorie", font=("Segoe UI", 12, "bold"), bg="lightgrey").grid(row=1, column=2, padx=10, pady=5,
                                                                                    sticky="w")
        self.category_var = StringVar()
        self.category_combobox = Combobox(inputs_frame, textvariable=self.category_var, values=categories,
                                          font=entry_font, state="readonly")
        self.category_combobox.current(0)
        self.category_combobox.grid(row=1, column=3, padx=10, pady=5, sticky="ew", ipady=5)


        Label(inputs_frame, text="Date", font=("Segoe UI", 12, "bold"), bg="lightgrey").grid(row=2, column=0, padx=10, pady=5,
                                                                               sticky="w")
        self.date_entry = DateEntry(inputs_frame, font=entry_font, date_pattern='yyyy-MM-dd',
                                    mindate=None, maxdate=None,
                                    year=datetime.now().year, month=datetime.now().month, day=datetime.now().day,
                                    state="normal")
        self.date_entry.grid(row=2, column=1, padx=10, pady=5, sticky="we", ipady=5)


        Label(inputs_frame, text="description", font=("Segoe UI", 12, "bold"), bg="lightgrey").grid(row=2, column=2, padx=10, pady=5,
                                                                               sticky="w")
        self.entry_description = Entry(inputs_frame, font=entry_font)
        self.entry_description.grid(row=2, column=3, padx=10, pady=5, sticky="ew", ipady=5)

        Button(inputs_frame, text="Ajouter Dépense", font=("Segoe UI", 12, "bold"),
               command=self.add_expense).grid(row=3, column=0, columnspan=4, pady=10, ipadx=25, ipady=8)


        style = Style()
        style.configure("Treeview", font=("Segoe UI", 12), rowheight=30)
        style.configure("Treeview.Heading", font=("Segoe UI", 14, "bold"))
        tree_frame = Frame(self.root)
        tree_frame.grid(row=2, column=0, columnspan=4, sticky="nsew", padx=10, pady=10)


        self.root.grid_rowconfigure(2, weight=1)
        self.root.grid_columnconfigure(0, weight=1)


        self.treeview = Treeview(tree_frame, columns=("Date", "Catégorie", "Montant", "Description"),
                                 show="headings", selectmode="browse")
        for col in ("Date", "Catégorie", "Montant", "Description"):
            self.treeview.heading(col, text=col, anchor="center")
            self.treeview.column(col, anchor="center", width=200, stretch=True, minwidth=100)
        self.treeview.grid(row=0, column=0, sticky="nsew")


        scrollbar_y = Scrollbar(tree_frame, orient=VERTICAL, command=self.treeview.yview)
        scrollbar_y.grid(row=0, column=1, sticky="ns", padx=10)
        self.treeview.configure(yscrollcommand=scrollbar_y.set)


        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        rows = load_expenses()
        self.manager.expenses = []

        for row in rows:
            if not row:
                continue
            exp_id, date_str, category, amount_str, description = row
            try:
                amount_float = float(amount_str)
            except:
                amount_float = 0.0
            expense = Expense(amount_float, category, date_str, exp_id, description)
            self.manager.add_expense(expense)
            self.treeview.insert("", "end", iid=exp_id,
                                 values=(date_str, category, f"{amount_float:.2f}", description))


        self.load_revenu()
        self.update_total()


        btn_frame = Frame(self.root, bg="white")
        btn_frame.grid(row=3, column=0, columnspan=4, sticky="ew", padx=10, pady=5)
        for i in range(4):
            btn_frame.grid_columnconfigure(i, weight=0)

        Button(btn_frame, text="Supprimer sélection", font=("Segoe UI", 12),
               command=self.delete_expense).grid(row=0, column=0, padx=15, pady=(5, 20), sticky="w")
        Button(btn_frame, text="Total Quotidien", font=("Segoe UI", 12),
               command=self.daily_total).grid(row=0, column=1, padx=15, pady=(5, 20), sticky="w")
        Button(btn_frame, text="Total Mensuel", font=("Segoe UI", 12),
               command=self.monthly_total).grid(row=0, column=2, padx=15, pady=(5, 20), sticky="w")

        Label(btn_frame, text="Total:", font=("Segoe UI", 12, "bold"), bg="white").grid(row=0, column=3, sticky="ew", padx=15, pady=(5, 20))
        Entry(btn_frame, textvariable=self.total_var, state="readonly", font=("Segoe UI", 12, "bold"), width=14).grid(row=0, column=4, sticky="e", padx=15, ipady=5, ipadx=5, pady=(5, 20))

        stats_frame = Frame(btn_frame, bg="white")
        stats_frame.grid(row=0, column=5, sticky="e")

        Button(stats_frame, text="Statistiques quotidiennes", font=("Segoe UI", 12),
               command=self.daily_statistics).pack(side=RIGHT, padx=10, pady=(5, 20))

        Button(stats_frame, text="Statistiques mensuelles", font=("Segoe UI", 12),
               command=self.monthly_statistics).pack(side=RIGHT, padx=10, pady=(5, 20))

        Button(btn_frame, text="Export to Excel", font=("Segoe UI", 12),
               command=self.export_to_excel).grid(row=0, column=6, padx=10, pady=(5, 20), sticky="e")

        def resize_columns(event):
            tree_width = self.treeview.winfo_width()
            num_cols = len(self.treeview["columns"])
            new_width = int(tree_width / num_cols) - 1
            for col in self.treeview["columns"]:
                self.treeview.column(col, width=new_width)
        self.root.bind("<Configure>", resize_columns)

        self.root.mainloop()

    def add_expense(self):
        montant = self.amount_entry.get()
        date = self.date_entry.get_date()
        categorie = self.category_combobox.get()
        description = self.entry_description.get()

        try:
            montant_float = float(montant)
            if montant_float <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre positif")
            return

        if not montant or not categorie:
            messagebox.showerror("Erreur", "Veuillez remplir tous les champs obligatoires")
            return

        expense_id = str(uuid.uuid4())
        expense = Expense(float(montant), categorie, date.strftime("%Y-%m-%d"), expense_id, description)
        self.manager.add_expense(expense)
        self.treeview.insert("", "end", iid=expense_id,
                             values=(date.strftime("%Y-%m-%d"), categorie, f"{float(montant):.2f}", description))
        self.update_total()

        rows = [[exp.id, exp.date, exp.category, f"{exp.amount:.2f}", exp.description] for exp in self.manager.expenses]
        save_expenses(rows)

    def delete_expense(self):
        selected_item = self.treeview.selection()
        if not selected_item:
            messagebox.showerror("Erreur", "Sélectionnez une dépense à supprimer")
            return
        expense_id = selected_item[0]
        try:
            self.manager.remove_expense(expense_id)
            self.treeview.delete(expense_id)
            self.update_total()
            rows = [[exp.id, exp.date, exp.category, f"{exp.amount:.2f}", exp.description] for exp in
                    self.manager.expenses]
            save_expenses(rows)


        except ValueError:
            messagebox.showerror("Erreur", "Dépense non trouvée")

    def update_total(self):
        total_depenses = sum(exp.amount for exp in self.manager.expenses)
        self.total_var.set(f"{total_depenses:.2f}")

        try:
            revenu = float(self.revenu_var.get())
        except ValueError:
            revenu = 0.0
        self.reste_var.set(f"{revenu - total_depenses:.2f}")

    def daily_total(self):
        popup = Toplevel(self.root)
        popup.title("Total Quotidien")
        popup.geometry("400x300")
        popup.resizable(False, False)

        Label(popup, text="Sélectionnez la date:", font=("Segoe UI", 14)).pack(pady=15)
        cal = DateEntry(popup, date_pattern='yyyy-MM-dd', font=("Segoe UI", 14))
        cal.pack(pady=10)

        daily_total_var = StringVar()
        Label(popup, text="Total:", font=("Segoe UI", 14)).pack(pady=10)
        Entry(popup, textvariable=daily_total_var, state="readonly", font=("Segoe UI", 14), width=12).pack(pady=5)

        def calculate_total():
            date_selected = cal.get()
            total = self.manager.daily_total(date_selected)
            daily_total_var.set(f"{total:.2f}")

        Button(popup, text="Calculer", font=("Segoe UI", 12), padx=15, pady=8, command=calculate_total).pack(pady=20)

    def monthly_total(self):
        popup = Toplevel(self.root)
        popup.title("Total Mensuel")
        popup.geometry("400x300")
        popup.resizable(False, False)

        Label(popup, text="Sélectionnez le mois:", font=("Segoe UI", 14)).pack(pady=5)
        month_cb = Combobox(popup, values=list(range(1, 13)), font=("Segoe UI", 14), state="readonly")
        month_cb.current(datetime.now().month - 1)
        month_cb.pack(pady=5)

        Label(popup, text="Sélectionnez l'année:", font=("Segoe UI", 14)).pack(pady=5)
        current_year = datetime.now().year
        year_cb = Combobox(popup, values=list(range(current_year - 5, current_year + 6)), font=("Segoe UI", 14),
                           state="readonly")
        year_cb.current(5)
        year_cb.pack(pady=5)

        total_label = Label(popup, text="Total: 0.00", font=("Segoe UI", 16, "bold"))
        total_label.pack(pady=20)

        def calculate_total():
            month = int(month_cb.get())
            year = int(year_cb.get())
            total = sum(exp.amount for exp in self.manager.expenses
                        if datetime.strptime(exp.date, "%Y-%m-%d").year == year
                        and datetime.strptime(exp.date, "%Y-%m-%d").month == month)
            total_label.config(text=f"Total: {total:.2f}")

        Button(popup, text="Calculer", font=("Segoe UI", 14, "bold"),
               width=20, command=calculate_total).pack(pady=10)

    def daily_statistics(self):
        popup = Toplevel(self.root)
        popup.title("Statistiques quotidiennes")
        popup.geometry("700x550")  # تكبير أكبر للنافذة
        popup.resizable(False, False)

        Label(popup, text="Sélectionnez la date:", font=("Segoe UI", 14)).pack(pady=10)
        cal = DateEntry(popup, date_pattern='yyyy-MM-dd', font=("Segoe UI", 14),  mindate=None, maxdate=None)
        cal.pack(pady=5)

        tree = Treeview(popup, columns=("Catégorie", "Montant", "Pourcentage"), show="headings")
        for col in ("Catégorie", "Montant", "Pourcentage"):
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=180)
        tree.pack(pady=10, padx=10, fill=BOTH, expand=True)


        total_label = Label(popup, text="Total: 0.00", font=("Segoe UI", 16, "bold"), fg="black")
        total_label.pack(pady=10)

        def calculate():
            date_selected = cal.get()
            categories_totals = {}
            for exp in self.manager.expenses:
                if exp.date == date_selected:
                    categories_totals[exp.category] = categories_totals.get(exp.category, 0) + exp.amount


            for i in tree.get_children():
                tree.delete(i)
            total = sum(categories_totals.values())
            for cat, amt in categories_totals.items():
                perc = (amt / total * 100) if total > 0 else 0
                tree.insert("", "end", values=(cat, f"{amt:.2f}", f"{perc:.1f}%"))
            total_label.config(text=f"Total: {total:.2f}")

        Button(popup, text="Calculer", font=("Segoe UI", 14), width=20, command=calculate).pack(pady=10)

    def monthly_statistics(self):
        popup = Toplevel(self.root)
        popup.title("Statistiques Mensuelles")
        popup.geometry("800x600")
        popup.resizable(True, True)

        Label(popup, text="Sélectionnez le mois:", font=("Segoe UI", 14)).pack(pady=5)
        month_cb = Combobox(popup, values=list(range(1, 13)), font=("Segoe UI", 14), state="readonly")
        month_cb.current(datetime.now().month - 1)
        month_cb.pack(pady=5)

        Label(popup, text="Sélectionnez l'année:", font=("Segoe UI", 14)).pack(pady=5)
        current_year = datetime.now().year
        year_cb = Combobox(popup, values=list(range(current_year - 5, current_year + 6)), font=("Segoe UI", 14),
                           state="readonly")
        year_cb.current(5)
        year_cb.pack(pady=5)

        tree = Treeview(popup, columns=("Catégorie", "Montant", "Pourcentage"), show="headings", height=8)
        for col in ("Catégorie", "Montant", "Pourcentage"):
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=200)
        tree.pack(pady=10, padx=10, fill=BOTH, expand=True)

        total_label = Label(popup, text="Total: 0.00", font=("Segoe UI", 16, "bold"))
        total_label.pack(pady=10)

        def calculate_statistics():
            month = int(month_cb.get())
            year = int(year_cb.get())
            categories_totals = {}
            for exp in self.manager.expenses:
                exp_date = datetime.strptime(exp.date, "%Y-%m-%d")
                if exp_date.year == year and exp_date.month == month:
                    categories_totals[exp.category] = categories_totals.get(exp.category, 0) + exp.amount

            for i in tree.get_children():
                tree.delete(i)

            total = sum(categories_totals.values())
            for cat, amt in categories_totals.items():
                perc = (amt / total * 100) if total > 0 else 0
                tree.insert("", "end", values=(cat, f"{amt:.2f}", f"{perc:.1f}%"))

            total_label.config(text=f"Total: {total:.2f}")


        btn_frame = Frame(popup)
        btn_frame.pack(pady=20)

        Button(btn_frame, text="Calculer", font=("Segoe UI", 14, "bold"), width=20, command=calculate_statistics).pack()

    def load_revenu(self):
        if os.path.exists("revenu.csv"):
            with open("revenu.csv", "r", encoding="utf-8") as f:
                contenu = f.read().strip().replace(",", ".")
                try:
                    value = float(contenu)
                except ValueError:
                    value = 0.0
                self.revenu_var.set(f"{value:.2f}")
        else:
            self.revenu_var.set("0.00")


    def save_revenu(self):
        raw = (self.revenu_var.get() or "").strip().replace(",", ".")
        try:
            value = float(raw)
        except ValueError:
            value = 0.0

        self.revenu_var.set(f"{value:.2f}")
        with open("revenu.csv", "w", encoding="utf-8") as f:
            f.write(f"{value:.2f}")

    def _normalize_and_save_revenu(self):
        self.save_revenu()
        self.update_total()

    def on_close(self):
        self.save_revenu()
        self.root.destroy()

    def export_to_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from tkinter.filedialog import asksaveasfilename


        file_path = asksaveasfilename(defaultextension=".xlsx",
                                      filetypes=[("Excel files", "*.xlsx")],
                                      title="Enregistrer le fichier Excel")
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Dépenses"


            columns = ["Date", "Catégorie", "Montant", "Description"]
            for col_num, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num, value=col_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")


            for row_num, exp in enumerate(self.manager.expenses, 2):
                ws.cell(row=row_num, column=1, value=exp.date)
                ws.cell(row=row_num, column=2, value=exp.category)
                ws.cell(row=row_num, column=3, value=exp.amount)
                ws.cell(row=row_num, column=4, value=exp.description)
                ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="right")


            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            wb.save(file_path)
            messagebox.showinfo("Succès", f"Les dépenses ont été exportées vers {file_path}")

        except Exception as e:
            messagebox.showerror("Erreur", f"Une erreur est survenue lors de l'export: {e}")


ExpenseApp()